import pandas as pd
from openpyxl import load_workbook, Workbook
from pathlib import Path

# путь к файлам
base_path = Path('\\\\pgk.rzd\\workgroups\\БД ЛП\\ЛПА (отдел аналитической работы)\\001_ШАБЛОНЫ\\111 Автоматический расчёт эффекта рынка\\Test\\')
source_file_name = 'мой.xlsx'
destination_file_name = 'мой_диспетчер.xlsx'
big_table_name = 'Bolbsha9_tablica.xlsx'
small_table_name = 'sovsem_malenbka9.xlsx'
medium_table_name = 'tablica_pomenbshe.xlsx'

# копируем данные из DataFrame в Excel-лист
def append_df_to_excel(workbook, sheet_name, df, start_row, start_col):
    # Если листа не существует, создаем его
    if sheet_name not in workbook.sheetnames:
        workbook.create_sheet(sheet_name)
    sheet = workbook[sheet_name]
    for i, row in df.iterrows():
        for j, value in enumerate(row):
            cell = sheet.cell(row=start_row + i, column=start_col + j)# + 1)
            cell.value = value

# Загрузка исходного файла или создание нового Workbook
workbook = load_workbook(base_path/source_file_name)

# Чтение большой таблицы
big_df = pd.read_excel(base_path/big_table_name, sheet_name='Sheet1')
# Выборка всех строк и столбцов от A до AR
big_df_subset = big_df.iloc[:, :42].copy()

# Выборка всех строк и столбцов от A до AR
big_df_subset2 = big_df.iloc[:, 42:43].copy()

# Выборка всех строк и столбцов от A до AR
big_df_subset3 = big_df.iloc[:, 43:44].copy()

# Копирование данных в лист 'СВОД Направления + фильтр'
append_df_to_excel(workbook, 'СВОД Направления + фильтр', big_df_subset, 7, 1)

# Копирование данных в лист 'СВОД Направления + фильтр'
append_df_to_excel(workbook, 'СВОД Направления + фильтр', big_df_subset2, 7, 86)

# Копирование данных в лист 'СВОД Направления + фильтр'
append_df_to_excel(workbook, 'СВОД Направления + фильтр', big_df_subset3, 7, 90)

# Сохранение изменений в новый файл
workbook.save(base_path/destination_file_name)

# Загрузка Workbook для дальнейших изменений
workbook = load_workbook(base_path/destination_file_name)

# Чтение и обработка маленькой таблицы
medium_df = pd.read_excel(base_path/medium_table_name, sheet_name='Sheet1')
# Выборка первых двух строк и столбцов от A до D
medium_df_subset = medium_df.iloc[:2, :4].copy()

# Копирование данных в лист 'выгрузка факт'
append_df_to_excel(workbook, 'выгрузка факт', medium_df_subset, 2, 1)

# Чтение и обработка средней таблицы
small_df = pd.read_excel(base_path/small_table_name, sheet_name='Sheet1')
# Выборка первых двух строк и столбцов от A до D
small_df_subset = small_df.iloc[:2, :4].copy()

# Копирование данных в лист 'выгрузка цель'
append_df_to_excel(workbook, 'выгрузка цель', small_df_subset, 2, 1)

# Сохранение всех изменений в 'мой_диспетчер.xlsx'
workbook.save('\\\\pgk.rzd\\workgroups\\БД ЛП\\ЛПА (отдел аналитической работы)\\001_ШАБЛОНЫ\\111 Автоматический расчёт эффекта рынка\\ПВ\\'+destination_file_name)
print('xlsx_test_PV.py - COMPLETE','\n','___________________________________________________________','\n')